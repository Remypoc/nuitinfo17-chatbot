"""
Load data from excel
"""
import unidecode
from openpyxl import load_workbook
from person import Personne

class DataLoader:

    def getAllPersons(self):
        persons = []
        filename = 'ressources/Data_GOT.xlsx'
        wb = load_workbook(filename)
        ws = wb.worksheets[0]
        for row in ws.iter_rows(row_offset=1, max_col=9, max_row=29):
            persons.append(
                Personne(self.format(row[0].value), self.format(row[1].value), self.format(row[2].value), self.format(row[3].value),
                         self.format(row[4].value), self.format(row[5].value), self.format(row[6].value),
                         self.format(row[7].value), self.format(row[8].value))
            )
        return persons

    def format(self, param):
        if type(param) is str:
            return unidecode.unidecode(param).replace("'", " ")
        return param

if __name__ == "__main__":
    d = DataLoader()
    persons = d.getAllPersons()
    for person in persons:
        print(person.pretty_print())
